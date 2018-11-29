#!/usr/bin/python
# Filename: cIO.py
try:
    import os, sys, logging
    sys.path.append(os.path.abspath(os.path.dirname(__file__)) + "\\openpyxl\\")
    import openpyxl as oxl
except:
    print("ExceptionERROR: Missing fundamental packages (required: os, sys, logging, openpyxl).")


class Read:
    def __init__(self, *args):
        # args[0] = full_file_name --  absolute path of a workbook
        # args[1] = worksheet -- INT sheet number in workbook
        # args[2] = logger

        try:
            self.open_wb(args[0], True, True)  # open workbook in read_only and data_only modes
        except:
            self.wb = ""

        try:
            self.open_ws(int(args[1]))
        except:
            try:
                self.open_ws(0)
            except:
                self.ws = ""

        try:
            self.logger = logging.getLogger(args[2])
        except:
            self.logger = logging.getLogger("logfile.log")

    def close_wb(self):
        try:
            self.wb.close()
        except:
            pass

    def col_increase_letter(self, letters, *step):
        # letters = STR - one or multiple letters of a column
        # step[0] = INT - number of letters to increase (optional, default = 1)
        # returns letter(s) / chr (alphabetically increased by step)
        try:
            step = step[0]
        except:
            step = 1
        chr_no = self.col_name_to_num(letters)
        chr_no += step
        return self.col_num_to_name(chr_no)

    def col_name_to_num(self, letters):
        pow = 1
        col_int = 0
        for letter in letters[::-1]:
            col_int += (int(letter, 36) - 9) * pow
            pow *= 26
        return col_int + 64

    def col_num_to_name(self, col_int):
        col_int -= 64
        letters = ''
        while col_int:
            mod = (col_int - 1) % 26
            letters += chr(mod + 65)
            col_int = (col_int - 1) // 26
        return ''.join(reversed(letters))

    def open_wb(self, xlsx_name, *read_modes):
        # read_modes[0] = read only -- BOOL (if true: read only = TRUE)
        # read_modes[1] = data only -- BOOL (if true: data only = TRUE)
        try:
            self.wb = oxl.load_workbook(filename=xlsx_name, read_only=read_modes[0], data_only=read_modes[1])
        except:
            try:
                self.wb = oxl.load_workbook(filename=xlsx_name)
            except:
                self.wb = ""
                self.logger.info("ERROR: Failed to access " + str(xlsx_name).split("\\")[-1] + ".")

    def open_ws(self, worksheet):
        try:
            self.ws = self.wb.worksheets[worksheet]
        except:
            self.ws = []
            self.logger.info("ERROR: No such worksheet available.")

    def read_cell(self, column, row):
        # column = CHR - cell column
        # row = INT - cell row
        # reads COLUMN / ROW cell
        try:
            cell_value = str(self.ws[str(column) + str(row)].value)
        except:
            cell_value = "None"
            self.logger.info("   * WARNING: Undefined cell " + str(column) + str(row))

        if not (cell_value == "None"):
            try:
                cell_value = float(cell_value)
            except:
                if cell_value.lower() == "inf":
                    cell_value = float(10 ** 10)
        return cell_value

    def read_column(self, column, start_row):
        # reads COLUMN beginning at START_ROW until it meets an empty cell
        # col = STR, e.g., col = "B"
        # start_row = INT
        # returns column as LIST
        self.logger.info("   * Reading data column (starting from " + str(column) + str(start_row) + ") ...")

        data = []
        valid_content = True
        __row__ = start_row
        while valid_content:
            try:
                cell_value = str(self.ws[str(column) + str(__row__)].value)
            except:
                cell_value = "None"
                self.logger.info("   * WARNING: Undefined cell " + str(column) + str(__row__))

            if not(cell_value == "None"):
                try:
                    data.append(float(cell_value))
                except:
                    if cell_value.lower() == "inf":
                        cell_value = float(10**10)
                    data.append(cell_value)
                __row__ += 1
            else:
                valid_content = False
        self.logger.info("   * OK")
        return data

    def read_row_str(self, row, start_col, **kwargs):
        # reads ROW beginning at COL until it meets an empty cell
        # row = INT
        # start_col = CHR, e.g., start_col = "B"

        # parse optional arguments
        try:
            for opt_var in kwargs.items():
                if "col_skip" in opt_var[0]:
                    # distance between columns with relevant data
                    col_skip = opt_var[1]
                if "end_col" in opt_var[0]:
                    # last relevant column
                    end_col = opt_var[1]
                if "if_row" in opt_var[0]:
                    # row that tells if the content of the actual cell is relevant
                    if_row = opt_var[1]
        except:
            pass

        if not ("col_skip" in locals()):
            col_skip = 1
        if not ("end_col" in locals()):
            end_col = "XFA"  # assumption: column XFA is the last column that can be handled within a spreadsheet
        if not ("if_row" in locals()):
            if_row = row

        self.logger.info(
            "   * Reading string row from " + str(self.xlsx_file).split("\\")[-1] + " (starting from " + str(start_col) + str(row) + ") ...")

        str_data = []
        valid_content = True
        __col__ = start_col
        end_col_num = self.col_name_to_num(end_col)

        while valid_content:
            cell_string = str(self.ws[str(__col__) + str(row)].value)
            valid_content = self.test_cell_content(str(__col__), if_row)
            if valid_content:
                str_data.append(cell_string)

            # update column to read next relevant row entry
            __col_num__ = self.col_name_to_num(__col__)  # convert ascii-chr to int
            __col_num__ += col_skip                      # add col_skip

            if not (__col_num__ > end_col_num):
                __col__ = self.col_num_to_name(__col_num__)  # re-convert int to ascii-chr
            else:
                valid_content = False
        self.logger.info("   * OK")
        return str_data

    def test_cell_content(self, column, row):
        # column = CHR - cell column
        # row = INT - cell row
        # reads COLUMN / ROW cell
        if str(self.ws[column + str(row)].value) == "None":
            return False
        else:
            return True

    def __call__(self):
        print("Class Info: <type> = XLSX manipulation in cIO.py")


class WorkbookContainer(Read):
    def __init__(self, *args):
        # args[0] = STR full_path to workbook template
        # args[1] = INT worksheet number
        # args[2] = STR logger
        try:
            Read.__init__(self, "#", "#", args[2])
        except:
            Read.__init__(self)

        try:
            self.open_wb(args[0])
            try:
                self.open_ws(args[1])
            except:
                pass
        except:
            pass

        # color definitions -- more: https://www.computerhope.com/htmcolor.htm
        self.white = 'FFFFFFFF'


    def save_close_wb(self, full_file_path):

        try:
            self.logger.info("   * Saving as: \n     " + full_file_path)
            self.wb.save(full_file_path)
            self.wb.close()
            self.logger.info("   * OK")
        except:
            self.logger.info("ERROR: Invalid file name or data.")

    def write_data_cell(self, column, row, value):
        # writes VALUE to cell
        try:
            self.ws[str(column) + str(row)].value = value
        except:
            self.logger.info("   * ERROR: Could not write value to CELL " + str(column) + str(row))

    def write_data_column(self, column, start_row, data_list):
        # writes COLUMN beginning at START_ROW
        # data_list is a LIST object
        self.logger.info("   * Writing column data (starting at " + str(column) + str(start_row) + ") ...")
        __row__ = start_row
        for val in data_list:
            self.ws[str(column) + str(__row__)].value = val
            __row__ += 1
        self.logger.info("   * OK")


